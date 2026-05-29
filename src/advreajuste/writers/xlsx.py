"""Geração da planilha pericial — formato alinhado à gold reference
`Reajuste_Elisangela.xls`.

Layout por vida (15 colunas):
  Ano | Mês | Valor pago | Reaj. aplic. | Idade | Tipo reaj. | Reaj. devido |
  Valor devido | Extra Pago | Extra devido | TOTAL PAGO | TOTAL DEVIDO |
  Diferença | Índ corr monet | Diferença corrigida

Cabeçalho do bloco:
  Plano de saúde | Categoria | Tipo
  Mês aniv. cont. | Vigência
  OBJETO: Reajuste por Sinistralidade
  Nome | Data Nasc.

Totalização geral ao final + resumo por índice de correção monetária.
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..calculators.substituicao import LinhaSubstituicao, TipoMes, totalizar_substituicao
from ..extractors.schemas import Beneficiario, Caso


FMT_BRL = 'R$ #,##0.00;[RED]-R$ #,##0.00'
FMT_PCT = "0.0000%"
FMT_IDX = "0.000000"

FILL_VERMELHO = PatternFill("solid", fgColor="FFC7CE")
FILL_VERDE = PatternFill("solid", fgColor="C6EFCE")
FILL_AZUL = PatternFill("solid", fgColor="DDEBF7")
FILL_CAB = PatternFill("solid", fgColor="305496")
FILL_SUBCAB = PatternFill("solid", fgColor="8EA9DB")
FILL_TOTAL = PatternFill("solid", fgColor="FFE699")

FONT_CAB = Font(color="FFFFFF", bold=True, size=11)
FONT_SUBCAB = Font(color="1F2937", bold=True, size=10)
FONT_VERMELHO = Font(color="9C0006")
FONT_VERDE = Font(color="006100")
FONT_AZUL = Font(color="1F4E78", bold=True)

BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)
CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
ESQ = Alignment(horizontal="left", vertical="center")

COLUNAS = [
    ("Ano", 6),
    ("Mês", 8),
    ("Valor pago", 14),
    ("Reaj. aplic.", 11),
    ("Idade", 7),
    ("Tipo reaj.", 13),
    ("Reaj. devido", 11),
    ("Valor devido", 14),
    ("Extra Pago", 11),
    ("Extra devido", 11),
    ("TOTAL PAGO", 14),
    ("TOTAL DEVIDO", 14),
    ("Diferença", 14),
    ("Índ corr monet", 13),
    ("Diferença corrigida", 16),
]

TIPO_LABEL = {
    TipoMes.NORMAL: "",
    TipoMes.ANIVERSARIO_ANS: "Anual (ANS ok)",
    TipoMes.ANIVERSARIO_ABUSIVO: "Anual ABUSIVO",
    TipoMes.ANIVERSARIO_DOWNGRADE: "Downgrade",
    TipoMes.PRO_RATA: "Pro-rata",
    TipoMes.ACERTO: "Acerto",
}


def _escrever_celula(ws, row, col, value, *, fill=None, font=None, fmt=None, border=True, align=CENTRO):
    c = ws.cell(row=row, column=col, value=value)
    if fill is not None:
        c.fill = fill
    if font is not None:
        c.font = font
    if fmt:
        c.number_format = fmt
    if border:
        c.border = BORDA
    c.alignment = align
    return c


def _escrever_cabecalho_bloco(ws, row: int, benef: Beneficiario, caso: Caso) -> int:
    """Escreve o cabeçalho de uma vida. Retorna o row onde começam os dados."""
    _escrever_celula(ws, row, 1, f"Plano de saúde: {caso.contrato.operadora}",
                     fill=FILL_SUBCAB, font=FONT_SUBCAB, align=ESQ)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    _escrever_celula(ws, row, 7, f"Categoria: {caso.contrato.tipo}",
                     fill=FILL_SUBCAB, font=FONT_SUBCAB, align=ESQ)
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=10)
    _escrever_celula(ws, row, 11, f"Apólice: {caso.contrato.numero}",
                     fill=FILL_SUBCAB, font=FONT_SUBCAB, align=ESQ)
    ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=15)

    row += 1
    _escrever_celula(ws, row, 1, f"Mês aniv. contrato: {caso.contrato.mes_aniversario:02d}",
                     fill=FILL_SUBCAB, font=FONT_SUBCAB, align=ESQ)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    _escrever_celula(ws, row, 7, f"Vigência: {caso.contrato.data_assinatura:%d/%m/%Y}",
                     fill=FILL_SUBCAB, font=FONT_SUBCAB, align=ESQ)
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=10)
    _escrever_celula(ws, row, 11, f"N.º vidas: {caso.contrato.n_vidas}",
                     fill=FILL_SUBCAB, font=FONT_SUBCAB, align=ESQ)
    ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=15)

    row += 1
    _escrever_celula(ws, row, 1, "OBJETO: Reajuste por Sinistralidade (substituição pelo teto ANS)",
                     fill=FILL_SUBCAB, font=FONT_SUBCAB, align=ESQ)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)

    row += 1
    _escrever_celula(ws, row, 1, f"Nome: {benef.nome}",
                     fill=FILL_CAB, font=FONT_CAB, align=ESQ)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _escrever_celula(ws, row, 9, f"Data Nasc.: {benef.data_nascimento:%d/%m/%Y}",
                     fill=FILL_CAB, font=FONT_CAB, align=ESQ)
    ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=15)

    row += 1
    for i, (h, _) in enumerate(COLUNAS, start=1):
        _escrever_celula(ws, row, i, h, fill=FILL_CAB, font=FONT_CAB)
    return row + 1


def _escrever_linhas_vida(
    ws, row_ini: int, linhas: list[LinhaSubstituicao],
    fator_por_comp: dict[str, Decimal] | None = None,
) -> int:
    """Escreve as linhas mensais. Retorna próxima row livre."""
    total_pago_acum = Decimal("0")
    total_devido_acum = Decimal("0")
    for i, l in enumerate(linhas):
        r = row_ini + i
        ano, mes = l.competencia.split("-")
        total_pago_acum += l.pago
        total_devido_acum += l.devido
        fator = fator_por_comp.get(l.competencia, Decimal("1")) if fator_por_comp else Decimal("1")
        diff_corr = l.diferenca * fator if l.diferenca > 0 else l.diferenca

        _escrever_celula(ws, r, 1, int(ano))
        _escrever_celula(ws, r, 2, int(mes))
        _escrever_celula(ws, r, 3, float(l.pago),
                         fill=FILL_VERMELHO, font=FONT_VERMELHO, fmt=FMT_BRL)
        _escrever_celula(ws, r, 4, float(l.reajuste_aplicado_pct), fmt=FMT_PCT)
        _escrever_celula(ws, r, 5, l.idade)
        _escrever_celula(ws, r, 6, TIPO_LABEL.get(l.tipo, ""), align=ESQ)
        _escrever_celula(ws, r, 7, float(l.reajuste_devido_pct), fmt=FMT_PCT)
        _escrever_celula(ws, r, 8, float(l.devido),
                         fill=FILL_VERDE, font=FONT_VERDE, fmt=FMT_BRL)
        _escrever_celula(ws, r, 9, 0, fmt=FMT_BRL)   # Extra Pago (coparticipação) — manual
        _escrever_celula(ws, r, 10, 0, fmt=FMT_BRL)  # Extra devido
        # Fórmulas vivas (TOTAL = valor + extra)
        col_pago = get_column_letter(3); col_extra_p = get_column_letter(9)
        col_dev = get_column_letter(8); col_extra_d = get_column_letter(10)
        f_tot_pago = f"={col_pago}{r}+{col_extra_p}{r}"
        f_tot_dev = f"={col_dev}{r}+{col_extra_d}{r}"
        _escrever_celula(ws, r, 11, f_tot_pago,
                         fill=FILL_VERMELHO, font=Font(color="9C0006", bold=True), fmt=FMT_BRL)
        _escrever_celula(ws, r, 12, f_tot_dev,
                         fill=FILL_VERDE, font=Font(color="006100", bold=True), fmt=FMT_BRL)
        f_diff = f"=K{r}-L{r}"
        _escrever_celula(ws, r, 13, f_diff,
                         fill=FILL_AZUL, font=FONT_AZUL, fmt=FMT_BRL)
        _escrever_celula(ws, r, 14, float(fator), fmt=FMT_IDX)
        _escrever_celula(ws, r, 15, float(diff_corr),
                         fill=FILL_AZUL, font=FONT_AZUL, fmt=FMT_BRL)

    # Linha de total da vida
    r_tot = row_ini + len(linhas)
    _escrever_celula(ws, r_tot, 1, "TOTAL", fill=FILL_TOTAL,
                     font=Font(bold=True), align=ESQ)
    ws.merge_cells(start_row=r_tot, start_column=1, end_row=r_tot, end_column=2)
    _escrever_celula(ws, r_tot, 11, f"=SUM(K{row_ini}:K{r_tot - 1})",
                     fill=FILL_TOTAL, font=Font(bold=True, color="9C0006"), fmt=FMT_BRL)
    _escrever_celula(ws, r_tot, 12, f"=SUM(L{row_ini}:L{r_tot - 1})",
                     fill=FILL_TOTAL, font=Font(bold=True, color="006100"), fmt=FMT_BRL)
    _escrever_celula(ws, r_tot, 13, f"=SUM(M{row_ini}:M{r_tot - 1})",
                     fill=FILL_TOTAL, font=Font(bold=True, color="1F4E78"), fmt=FMT_BRL)
    _escrever_celula(ws, r_tot, 15, f"=SUM(O{row_ini}:O{r_tot - 1})",
                     fill=FILL_TOTAL, font=Font(bold=True, color="1F4E78"), fmt=FMT_BRL)
    return r_tot + 2


def _aba_resumo_correcao(wb, correcao: dict, data_alvo: date, totais_por_benef: list[dict]):
    ws = wb.create_sheet("Correção Monetária")
    ws["A1"] = f"Correção monetária — atualização até {data_alvo:%d/%m/%Y}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Diferença histórica aplicada sobre cada mês com reajuste abusivo."
    ws["A2"].font = Font(italic=True, color="595959")

    headers = ["Índice", "Total atualizado", "Observação"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = FONT_CAB; c.fill = FILL_CAB; c.alignment = CENTRO

    r = 5
    valor_hist = correcao["diferenca_historica"]
    _escrever_celula(ws, r, 1, "Histórico (sem correção)", align=ESQ)
    _escrever_celula(ws, r, 2, float(valor_hist), fmt=FMT_BRL)
    _escrever_celula(ws, r, 3, "Soma bruta das diferenças mensais", align=ESQ)
    r += 1
    for ind, v in correcao["totais_por_indice"].items():
        erros = correcao["erros_por_indice"].get(ind, 0)
        obs = f"{erros} mês(es) sem dado BACEN" if erros else "OK"
        _escrever_celula(ws, r, 1, ind, align=ESQ)
        _escrever_celula(ws, r, 2, float(v), fmt=FMT_BRL)
        _escrever_celula(ws, r, 3, obs, align=ESQ)
        r += 1
    _escrever_celula(ws, r, 1, "Juros 1% a.m. (pré-Lei 14.905/24)", align=ESQ)
    _escrever_celula(ws, r, 2, float(correcao["juros_1pct_am"]), fmt=FMT_BRL)
    _escrever_celula(ws, r, 3, "Juros simples", align=ESQ)
    r += 1
    for label, v in correcao["combinacoes"].items():
        _escrever_celula(ws, r, 1, label, align=ESQ)
        _escrever_celula(ws, r, 2, float(v), fmt=FMT_BRL,
                         fill=FILL_AZUL, font=FONT_AZUL)
        _escrever_celula(ws, r, 3, "Combinação usual na prática forense", align=ESQ)
        r += 1

    # Resumo por beneficiário
    r += 2
    c = ws.cell(row=r, column=1, value="Resumo por beneficiário")
    c.font = Font(bold=True, size=12); r += 1
    hs = ["Beneficiário", "Pago", "Devido", "Diferença histórica",
          "Restituível trienal simples", "Em dobro (CDC 42)",
          "Reajustes abusivos", "Meses"]
    for i, h in enumerate(hs, 1):
        c = ws.cell(row=r, column=i, value=h); c.font = FONT_CAB; c.fill = FILL_CAB; c.alignment = CENTRO
    r += 1
    for t in totais_por_benef:
        ws.cell(row=r, column=1, value=t["nome"]).alignment = ESQ
        for i, k in enumerate(
            ["total_pago", "total_devido", "diferenca",
             "restituivel_simples", "restituivel_dobro_art42",
             "n_aniversarios_abusivos", "n_meses"], start=2
        ):
            cell = ws.cell(row=r, column=i, value=float(t[k]) if isinstance(t[k], Decimal) else t[k])
            if i <= 6:
                cell.number_format = FMT_BRL
        r += 1

    for col, w in [("A", 38), ("B", 18), ("C", 30), ("D", 18), ("E", 22), ("F", 20), ("G", 18), ("H", 10)]:
        ws.column_dimensions[col].width = w


def _calcular_fator_por_mes(correcao: dict | None, indice_principal: str) -> dict[str, Decimal]:
    """Para cada competência, retorna fator de correção desde aquele mês até data alvo."""
    if not correcao:
        return {}
    from ..calculators.correcao_monetaria import fator_acumulado

    data_alvo = date.fromisoformat(correcao["data_alvo"])
    fatores: dict[str, Decimal] = {}
    for comp in correcao["diferenca_por_mes"].keys():
        y, m = (int(x) for x in comp.split("-"))
        try:
            if indice_principal == "TJSP":
                # usar INPC até marco 14905, depois SELIC/IPCA
                from ..calculators.correcao_monetaria import MARCO_LEI_14905
                if data_alvo <= MARCO_LEI_14905:
                    fatores[comp] = fator_acumulado("INPC", date(y, m, 1), data_alvo)
                else:
                    fatores[comp] = fator_acumulado("INPC", date(y, m, 1), min(data_alvo, MARCO_LEI_14905))
            else:
                fatores[comp] = fator_acumulado(indice_principal, date(y, m, 1), data_alvo)
        except Exception:
            fatores[comp] = Decimal("1")
    return fatores


def gerar_planilha_pericial(
    saida: Path,
    caso: Caso,
    resultados: dict[Beneficiario, list[LinhaSubstituicao]],
    correcao: dict | None = None,
    indice_principal: str = "TJSP",
    data_distribuicao: date | None = None,
) -> Path:
    """Gera XLSX no formato pericial (uma aba por vida + aba de correção)."""
    wb = Workbook()
    wb.remove(wb.active)

    # Calcular fator de correção por competência para coluna "Diferença corrigida"
    fator_por_comp = _calcular_fator_por_mes(correcao, indice_principal)

    totais_por_benef: list[dict] = []
    for i, (benef, linhas) in enumerate(resultados.items(), start=1):
        sheet_name = f"{i:02d}-{benef.nome.split()[0][:20]}"
        ws = wb.create_sheet(sheet_name[:31])
        for col, (_, w) in enumerate(COLUNAS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w
        ws["A1"] = f"DEMONSTRATIVO DE CÁLCULO — {caso.caso_id}"
        ws["A1"].font = Font(bold=True, size=13)
        ws.merge_cells("A1:O1")
        row = _escrever_cabecalho_bloco(ws, 3, benef, caso)
        _escrever_linhas_vida(ws, row, linhas, fator_por_comp)
        ws.freeze_panes = f"A{row + 1}"

        tot = totalizar_substituicao(linhas)
        totais_por_benef.append({"nome": benef.nome, **tot})

    # Aba resumo
    if correcao:
        _aba_resumo_correcao(wb, correcao, data_distribuicao or date.today(), totais_por_benef)
    else:
        ws = wb.create_sheet("Resumo")
        ws["A1"] = "Resumo geral"
        ws["A1"].font = Font(bold=True, size=14)
        r = 3
        hs = ["Beneficiário", "Pago", "Devido", "Diferença", "Restituível trienal", "Abusivos"]
        for i, h in enumerate(hs, 1):
            c = ws.cell(row=r, column=i, value=h); c.font = FONT_CAB; c.fill = FILL_CAB
        r += 1
        for t in totais_por_benef:
            ws.cell(row=r, column=1, value=t["nome"])
            for i, k in enumerate(
                ["total_pago", "total_devido", "diferenca",
                 "restituivel_simples", "n_aniversarios_abusivos"], start=2
            ):
                cell = ws.cell(row=r, column=i, value=float(t[k]) if isinstance(t[k], Decimal) else t[k])
                if i <= 5:
                    cell.number_format = FMT_BRL
            r += 1
        for col, w in [("A", 35), ("B", 16), ("C", 16), ("D", 16), ("E", 20), ("F", 12)]:
            ws.column_dimensions[col].width = w

    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)
    return saida


# Alias compatibilidade com código existente
def gerar_planilha(saida, resultados_map, caso_id, total_correcao=None):
    """Deprecated — usa gerar_planilha_pericial. Mantido para compat com pipeline antigo."""
    wb = Workbook()
    ws = wb.active
    ws.title = str(caso_id)[:31]
    ws["A1"] = f"Memória de cálculo — caso {caso_id}"
    ws["A1"].font = Font(bold=True, size=14)
    r = 4
    for b, linhas in resultados_map.items():
        ws.cell(row=r, column=1, value=b.nome).font = Font(bold=True)
        r += 1
        for h in ["Competência", "Idade", "Pago", "Devido", "Diferença"]:
            ws.cell(row=r, column=["Competência", "Idade", "Pago", "Devido", "Diferença"].index(h) + 1, value=h)
        r += 1
        for l in linhas:
            ws.cell(row=r, column=1, value=l.competencia)
            ws.cell(row=r, column=2, value=l.idade if hasattr(l, "idade") else "")
            val_pago = float(getattr(l, "cobrada", None) or getattr(l, "pago", 0))
            val_dev = float(getattr(l, "devida", None) or getattr(l, "devido", 0))
            val_diff = float(getattr(l, "delta", None) or getattr(l, "diferenca", 0))
            ws.cell(row=r, column=3, value=val_pago).number_format = FMT_BRL
            ws.cell(row=r, column=4, value=val_dev).number_format = FMT_BRL
            ws.cell(row=r, column=5, value=val_diff).number_format = FMT_BRL
            r += 1
        r += 1
    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)
    return saida
