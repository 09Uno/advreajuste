"""Writer XLSX usando o TEMPLATE OFICIAL do escritório do mentorando.

Preenche as 13 abas do template `templates/planilha_modelo_mentorando.xlsx`:
  - Calculo{N}Vidas: aba principal com tabela 2004-2025 mês a mês
  - Retroativo: cálculo de retroativo parcelado
  - % Acumulada: tabela ANS vs Operadora (estática + ano corrente)
  - Relatorio: relatório de output formatado
  - Honorarios: proposta de honorários por fase
  - CalculoMensalidade: mensalidade total devida atual
  - Projecao: projeção futura
  - IndicesANSIndiv: índices ANS individuais (estático)

A estratégia é COPIAR o template (preservando fórmulas, formatação,
abas estáticas) e PREENCHER apenas células dinâmicas baseadas no caso.
"""
from __future__ import annotations

import shutil
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from ..calculators.ans import ANS_INDIVIDUAL
from ..calculators.faixa_etaria import idade_em
from ..calculators.substituicao import LinhaSubstituicao, TipoMes, totalizar_substituicao
from ..extractors.schemas import Beneficiario, Caso


def _descobrir_template() -> Path:
    """Descobre o template em vários paths possíveis (local + Streamlit Cloud)."""
    aqui = Path(__file__).resolve()
    candidatos = [
        # src/advreajuste/writers/ → ../../../templates/
        aqui.parent.parent.parent.parent / "templates" / "planilha_modelo_mentorando.xlsx",
        # /mount/src/REPO/templates/ no Streamlit Cloud
        aqui.parent.parent.parent / "templates" / "planilha_modelo_mentorando.xlsx",
        # CWD/templates/
        Path.cwd() / "templates" / "planilha_modelo_mentorando.xlsx",
        # Hard-coded fallback
        Path("templates/planilha_modelo_mentorando.xlsx"),
    ]
    for c in candidatos:
        if c.exists():
            return c
    # Retorna o primeiro candidato pra exibir mensagem de erro útil
    return candidatos[0]


TEMPLATE_PATH = _descobrir_template()

MESES_PT = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
            "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

CATEGORIA_LABEL = {
    "individual": "Individual",
    "familiar": "Familiar",
    "coletivo_empresarial": "Coletivo Empresarial",
    "coletivo_adesao": "Coletivo por Adesão",
}

TIPO_REAJ_LABEL = {
    TipoMes.NORMAL: "",
    TipoMes.ANIVERSARIO_ANS: "Anual",
    TipoMes.ANIVERSARIO_ABUSIVO: "Anual",
    TipoMes.ANIVERSARIO_DOWNGRADE: "Downgrade",
    TipoMes.PRO_RATA: "Pro-rata",
    TipoMes.ACERTO: "Acerto",
}


LINHA_DADOS_INI = 13
LINHA_DADOS_FIM = 276
LINHA_TOTAL_DIFERENCA = 277


def _life_starts(n_vidas: int) -> list[int]:
    """Colunas iniciais dos blocos de cada vida no template."""
    if n_vidas <= 1:
        return [3]  # C
    return [4 + i * 11 for i in range(min(n_vidas, 6))]  # D, O, Z, AK...


def _cols_vida(n_vidas: int, benef_idx: int) -> dict[str, int]:
    """Mapa real das colunas de uma vida na aba Calculo{N}Vida(s)."""
    start = _life_starts(n_vidas)[benef_idx]
    return {
        "valor_pago": start,
        "reaj_aplic": start + 1,
        "idade": start + 2,
        "tipo_reaj": start + 3,
        "reaj_devido": start + 4,
        "valor_devido": start + 5,
        "extra_pago": start + 6,
        "extra_devido": start + 7,
        "total_pago": start + 8,
        "total_devido": start + 9,
        "diferenca": start + 10,
    }


def _cols_total(n_vidas: int) -> dict[str, int]:
    """Mapa das colunas totalizadoras da aba usada."""
    if n_vidas <= 1:
        return {
            "total_pago": 11,      # K
            "total_devido": 12,    # L
            "diferenca": 13,       # M
            "dif_corrigida": 15,   # O
        }
    start = 4 + min(n_vidas, 6) * 11
    return {
        "total_pago": start + 2,
        "total_devido": start + 4,
        "diferenca": start + 6,
        "dif_corrigida": start + 8,
    }


def _aba_calculo_n(n_vidas: int) -> str:
    """Retorna nome da aba CalculoNVidas correta (limita a 6)."""
    n = min(max(n_vidas, 1), 6)
    return f"Calculo{n}Vida" if n == 1 else f"Calculo{n}Vidas"


def _linha_para_mes(ano: int, mes: int) -> int:
    """Mapeia (ano, mês) → linha na tabela (L13 = 2004-01)."""
    return 13 + (ano - 2004) * 12 + (mes - 1)


def _preencher_cabecalho(ws, caso: Caso, n_vidas: int):
    """Preenche cabeçalho da aba CalculoNVidas (linhas 1-11)."""
    contrato = caso.contrato
    cat_label = CATEGORIA_LABEL.get(contrato.tipo, contrato.tipo)

    if n_vidas <= 1:
        _safe_set(ws, 6, 3, contrato.operadora or "—")  # C6
        _safe_set(ws, 6, 7, cat_label)                  # G6
        _safe_set(ws, 6, 14, cat_label)                 # N6
        _safe_set(ws, 7, 3, MESES_PT[contrato.mes_aniversario - 1])  # C7
        _safe_set(ws, 7, 7, contrato.data_assinatura)   # G7
    else:
        _safe_set(ws, 6, 5, contrato.operadora or "—")  # E6
        _safe_set(ws, 6, 10, cat_label)                 # J6
        _safe_set(ws, 6, 14, cat_label)                 # N6
        _safe_set(ws, 7, 5, MESES_PT[contrato.mes_aniversario - 1])  # E7
        _safe_set(ws, 7, 10, contrato.data_assinatura)  # J7

    # OBJETO já vem preenchido no template

    benefs = list(caso.beneficiarios)[:n_vidas]
    for idx, b in enumerate(benefs):
        if n_vidas <= 1:
            nome_col, nasc_col = 3, 9  # C11 / I11
        else:
            start = _life_starts(n_vidas)[idx]
            nome_col, nasc_col = start + 1, start + 6
        _safe_set(ws, 11, nome_col, b.nome)
        _safe_set(
            ws,
            11,
            nasc_col,
            b.data_nascimento if b.data_nascimento != date(1970, 1, 1) else "—",
        )


def _preencher_linhas_vida(
    ws,
    beneficiario: Beneficiario,
    benef_idx: int,
    linhas: list[LinhaSubstituicao],
    n_vidas: int,
):
    """Preenche linhas mensais da vida `benef_idx` (0-based).

    O template tem formulas e totais ao lado de cada bloco. Para adaptar o
    modelo sem contaminar com dados de exemplo, preenchemos todos os meses
    2004-2025 com a ultima mensalidade conhecida e mantemos diferenca zero
    fora do periodo calculado.
    """
    if not linhas:
        return

    cols = _cols_vida(n_vidas, benef_idx)
    por_comp = {l.competencia: l for l in linhas}
    linhas_ord = sorted(linhas, key=lambda l: l.competencia)
    primeiro = linhas_ord[0]
    ultimo_conhecido = primeiro

    for ano in range(2004, 2026):
        for mes in range(1, 13):
            comp = f"{ano:04d}-{mes:02d}"
            row = _linha_para_mes(ano, mes)
            l = por_comp.get(comp)
            dentro_periodo = l is not None
            if l is None:
                if comp < primeiro.competencia:
                    l = primeiro
                    pago = devido = primeiro.pago
                    reaj_aplic = reaj_devido = Decimal("0")
                    tipo_label = ""
                    diff = Decimal("0")
                else:
                    l = ultimo_conhecido
                    pago = ultimo_conhecido.pago
                    devido = ultimo_conhecido.devido
                    reaj_aplic = reaj_devido = Decimal("0")
                    tipo_label = ""
                    diff = ultimo_conhecido.diferenca
            else:
                ultimo_conhecido = l
                pago = l.pago
                devido = l.devido
                reaj_aplic = l.reajuste_aplicado_pct
                reaj_devido = l.reajuste_devido_pct
                tipo_label = TIPO_REAJ_LABEL.get(l.tipo, "") if l.tipo != TipoMes.NORMAL else ""
                diff = l.diferenca

            _safe_set(ws, row, cols["valor_pago"], float(pago))
            _safe_set(ws, row, cols["reaj_aplic"], float(reaj_aplic) if dentro_periodo else 0)
            _safe_set(ws, row, cols["idade"], idade_em(beneficiario.data_nascimento, date(ano, mes, 1)))
            _safe_set(ws, row, cols["tipo_reaj"], tipo_label)
            _safe_set(ws, row, cols["reaj_devido"], float(reaj_devido) if dentro_periodo else 0)
            _safe_set(ws, row, cols["valor_devido"], float(devido))
            _safe_set(ws, row, cols["extra_pago"], 0)
            _safe_set(ws, row, cols["extra_devido"], 0)


def _preencher_linhas_neutras(ws, benef_idx: int, n_vidas: int):
    """Preenche um bloco oculto com valores neutros para evitar #DIV/0!."""
    cols = _cols_vida(n_vidas, benef_idx)
    for ano in range(2004, 2026):
        for mes in range(1, 13):
            row = _linha_para_mes(ano, mes)
            _safe_set(ws, row, cols["valor_pago"], 1)
            _safe_set(ws, row, cols["reaj_aplic"], 0)
            _safe_set(ws, row, cols["idade"], "")
            _safe_set(ws, row, cols["tipo_reaj"], "")
            _safe_set(ws, row, cols["reaj_devido"], 0)
            _safe_set(ws, row, cols["valor_devido"], 1)
            _safe_set(ws, row, cols["extra_pago"], 0)
            _safe_set(ws, row, cols["extra_devido"], 0)


def _preencher_total_auxiliar(ws, n_vidas: int):
    """Preenche coluna auxiliar de total pago nas abas multi-vidas.

    O template tem um segundo mini-bloco "Vl. tot. pago / Reaj. aplic." ao
    fim das abas com 2+ vidas. A coluna de valor vem vazia no arquivo modelo,
    mas a coluna de reajuste divide por ela; preencher a referencia evita
    #DIV/0! quando o Excel recalcula o workbook.
    """
    if n_vidas <= 1:
        return
    total_cols = _cols_total(n_vidas)
    aux_col = 4 + n_vidas * 11 + 10
    total_pago_letter = get_column_letter(total_cols["total_pago"])
    for row in range(LINHA_DADOS_INI, LINHA_DADOS_FIM + 1):
        _safe_set(ws, row, aux_col, f"={total_pago_letter}{row}")


def _preencher_relatorio(ws, caso: Caso, resultados, correcao: dict | None):
    """Preenche aba Relatorio com dados do caso."""
    contrato = caso.contrato
    cat_label = CATEGORIA_LABEL.get(contrato.tipo, contrato.tipo)

    # L4: OPERADORA / CATEGORIA / TIPO
    for r in range(1, 50):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            v_lower = v.strip().lower()
            if v_lower == "operadora":
                _safe_set(ws, r, c + 1, contrato.operadora or "—")
            elif v_lower == "categoria":
                _safe_set(ws, r, c + 1, cat_label)
            elif v_lower == "tipo":
                _safe_set(ws, r, c + 1, cat_label)
            elif v_lower.startswith("mês aniv"):
                _safe_set(ws, r, c + 1, MESES_PT[contrato.mes_aniversario - 1])
            elif v_lower == "vigência":
                _safe_set(ws, r, c + 1, contrato.data_assinatura)

    # Preenche nomes dos beneficiários (até o nº disponível)
    benefs = list(caso.beneficiarios)
    nome_idx = 0
    for r in range(10, 18):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() == "NOME" and nome_idx < len(benefs):
                b = benefs[nome_idx]
                _safe_set(ws, r, c + 1, b.nome)
                # Data nascimento na mesma linha
                for c2 in range(c + 1, ws.max_column + 1):
                    v2 = ws.cell(r, c2).value
                    if isinstance(v2, str) and "DATA NASCIMENTO" in v2.upper():
                        if b.data_nascimento != date(1970, 1, 1):
                            _safe_set(ws, r, c2 + 1, b.data_nascimento)
                        else:
                            _safe_set(ws, r, c2 + 1, "—")
                        break
                nome_idx += 1
                break


def _ultima_linha_util(linhas: list[LinhaSubstituicao]) -> int:
    """Ultima linha do periodo calculado dentro do range fixo do template."""
    rows = []
    for l in linhas:
        ano, mes = (int(x) for x in l.competencia.split("-"))
        row = _linha_para_mes(ano, mes)
        if LINHA_DADOS_INI <= row <= LINHA_DADOS_FIM:
            rows.append(row)
    return max(rows, default=LINHA_DADOS_FIM)


def _ref(aba: str, col: int, row: int) -> str:
    return f"'{aba}'!{get_column_letter(col)}{row}"


def _preencher_formulas_relatorio(
    ws,
    caso: Caso,
    aba_calc: str,
    n_vidas: int,
    ultima_linha: int,
):
    """Atualiza o Relatorio para apontar para a aba de calculo realmente usada.

    O template original mistura referencias a Calculo1Vida e Calculo6Vidas.
    Quando o caso tem outra quantidade de vidas, essas formulas ficam
    desalinhadas. Aqui mantemos o layout do relatorio, mas redirecionamos as
    formulas para os blocos corretos.
    """
    total_cols = _cols_total(n_vidas)

    _safe_set(ws, 91, 2, f"={_ref(aba_calc, total_cols['total_pago'], ultima_linha)}")
    _safe_set(ws, 91, 6, f"={_ref(aba_calc, total_cols['total_devido'], ultima_linha)}")
    _safe_set(ws, 96, 5, "=B91-F91")
    _safe_set(ws, 97, 5, "=IFERROR((F91/B91)-1,0)")
    _safe_set(ws, 100, 4, f"={_ref(aba_calc, total_cols['diferenca'], LINHA_TOTAL_DIFERENCA)}")

    benefs = list(caso.beneficiarios)[:n_vidas]
    for idx in range(6):
        row = 107 + idx
        if idx < len(benefs):
            cols = _cols_vida(n_vidas, idx)
            _safe_set(ws, row, 2, f"=C{10 + idx}")
            _safe_set(ws, row, 4, f"={_ref(aba_calc, cols['total_pago'], ultima_linha)}")
            _safe_set(ws, row, 5, f"={_ref(aba_calc, cols['total_devido'], ultima_linha)}")
            _safe_set(ws, row, 6, f"=D{row}-E{row}")
            _safe_set(ws, row, 7, f"={_ref(aba_calc, cols['diferenca'], LINHA_TOTAL_DIFERENCA)}")
        else:
            _safe_set(ws, row, 2, "")
            _safe_set(ws, row, 4, "")
            _safe_set(ws, row, 5, "")
            _safe_set(ws, row, 6, "")
            _safe_set(ws, row, 7, "")


def _preencher_honorarios(ws, caso: Caso, total_restituicao: Decimal):
    """Preenche REQUERENTE/REQUERIDO + valor referência."""
    nomes = ", ".join(b.nome for b in caso.beneficiarios)
    for r in range(1, 20):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            v_strip = v.strip().upper()
            if v_strip == "REQUERENTE":
                _safe_set(ws, r, c + 1, nomes)
            elif v_strip == "REQUERIDO":
                _safe_set(ws, r, c + 1, caso.contrato.operadora or "—")


def _linhas_por_ano(linhas: list[LinhaSubstituicao]) -> dict[int, list[LinhaSubstituicao]]:
    """Agrupa linhas por ano."""
    out: dict[int, list[LinhaSubstituicao]] = {}
    for l in linhas:
        ano = int(l.competencia.split("-")[0])
        out.setdefault(ano, []).append(l)
    return out


def _anos_abusivos(linhas: list[LinhaSubstituicao]) -> list[tuple[int, LinhaSubstituicao]]:
    """Retorna [(ano, linha_aniversario_abusiva)] em ordem cronologica."""
    out = []
    for l in linhas:
        if l.tipo == TipoMes.ANIVERSARIO_ABUSIVO:
            out.append((int(l.competencia.split("-")[0]), l))
    return out


def _safe_set(ws, r: int, c: int, value):
    """Seta value em uma celula, pulando MergedCells (read-only)."""
    cell = ws.cell(r, c)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _limpar_intervalo(ws, r_ini: int, r_fim: int, c_ini: int, c_fim: int):
    """Limpa values em um retangulo de celulas (mantem formatacao e merges)."""
    for r in range(r_ini, r_fim + 1):
        for c in range(c_ini, c_fim + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _preencher_bloco_retroativo(
    ws,
    linha_base: int,
    ano: int,
    mes_aniv: int,
    valor_base: Decimal,
    pct_aplicado: Decimal,
    pct_ans: Decimal,
):
    """Preenche um bloco de 19 linhas com dados de retroativo de um ano.

    Layout (linhas relativas a linha_base):
      +0 (L1): Titulo "Calculadora reajuste retroativo {ano}"
      +4 (L5): Headers (Valor antes / Reaj aplic / Valor c/reaj | mes | pago | %retro | devido | val c/retro | retroativo)
      +5 (L6): Dados linha 1 + tabela mensal mes 1
      +6 (L7): Headers (Valor retro / Meses retro / Total retro)
      +7 (L8): Dados linha 2
      +8 (L9): Headers (Qtd parcelas / Valor parcela / % retro / % retro + anual)
      +9 (L10): Dados linha 3
      +10 (L11): "Total a ser pago", _, total
      +11..+16 (L12-L17): tabela mensal meses 7-12
      +17 (L18): "Tota Retroativo a ser pago", _, total
      +18 (L19): "Retroativo Parcelado em 12", _, parcela
    """
    # Limpa todas as celulas do bloco (19 linhas, 10 cols)
    _limpar_intervalo(ws, linha_base, linha_base + 18, 1, 10)

    valor_com_reaj = valor_base * (Decimal(1) + pct_aplicado)
    valor_devido = valor_base * (Decimal(1) + pct_ans)
    retro_unit = valor_com_reaj - valor_devido
    meses_retro = 12 - mes_aniv + 1  # do mes aniv até dezembro
    if meses_retro < 0:
        meses_retro = 0
    total_retro = retro_unit * Decimal(meses_retro)
    valor_parcela = (total_retro / Decimal(12)) if total_retro else Decimal(0)
    pct_retro = (retro_unit / valor_devido) if valor_devido else Decimal(0)
    pct_retro_anual = pct_retro + pct_aplicado

    # Titulo
    _safe_set(ws, linha_base, 1, f"Calculadora reajuste retroativo {ano}")

    # Headers L5 - bloco esquerdo
    L5 = linha_base + 4
    _safe_set(ws, L5, 1, "Valor antes do reajuste")
    _safe_set(ws, L5, 2, "Reajuste aplicado ")
    _safe_set(ws, L5, 3, "Valor com reajuste ")
    # Headers L5 - tabela mensal
    _safe_set(ws, L5, 5, "Mês")
    _safe_set(ws, L5, 6, f"Valor Pago em {ano}")
    _safe_set(ws, L5, 7, "% retro")
    _safe_set(ws, L5, 8, f"Valor Devido em {ano}")
    _safe_set(ws, L5, 9, "Valor Devido (com retro)")
    _safe_set(ws, L5, 10, "Retroativo")

    # L6 - dados linha 1
    L6 = linha_base + 5
    _safe_set(ws, L6, 1, float(valor_base))
    _safe_set(ws, L6, 2, float(pct_aplicado))
    _safe_set(ws, L6, 3, float(valor_com_reaj))

    # L7 - headers linha 2
    L7 = linha_base + 6
    _safe_set(ws, L7, 1, "Valor do retroativo")
    _safe_set(ws, L7, 2, "Meses retroativos ")
    _safe_set(ws, L7, 3, "Total retroativo ")

    # L8 - dados linha 2
    L8 = linha_base + 7
    _safe_set(ws, L8, 1, float(retro_unit))
    _safe_set(ws, L8, 2, meses_retro)
    _safe_set(ws, L8, 3, float(total_retro))

    # L9 - headers linha 3
    L9 = linha_base + 8
    _safe_set(ws, L9, 1, "Quantidade de parcelas ")
    _safe_set(ws, L9, 2, "Valor da parcela")
    _safe_set(ws, L9, 3, "% retroativo")
    _safe_set(ws, L9, 4, "% retroativo + % anual")

    # L10 - dados linha 3
    L10 = linha_base + 9
    _safe_set(ws, L10, 1, 12)
    _safe_set(ws, L10, 2, float(valor_parcela))
    _safe_set(ws, L10, 3, float(pct_retro))
    _safe_set(ws, L10, 4, float(pct_retro_anual))

    # L11 - total a ser pago
    L11 = linha_base + 10
    _safe_set(ws, L11, 1, "Total a ser pago ")
    _safe_set(ws, L11, 3, float(valor_base * Decimal(12) + total_retro))

    # Tabela mensal: meses 1-12 nas linhas L6..L17
    for i in range(12):
        mes = i + 1
        row = linha_base + 5 + i  # L6 = mes 1
        _safe_set(ws, row, 5, mes)
        if mes < mes_aniv:
            # Antes do aniversario: pagou base, devia base
            _safe_set(ws, row, 6, float(valor_base))
            _safe_set(ws, row, 7, 0)
            _safe_set(ws, row, 8, 0)
            _safe_set(ws, row, 9, float(valor_base))
            _safe_set(ws, row, 10, 0)
        elif mes == mes_aniv:
            # Mes aniversario: aplicou reajuste vs ANS
            _safe_set(ws, row, 6, float(valor_com_reaj))
            _safe_set(ws, row, 7, float(pct_aplicado))
            _safe_set(ws, row, 8, float(pct_ans))
            _safe_set(ws, row, 9, float(valor_devido))
            _safe_set(ws, row, 10, float(retro_unit))
        else:
            # Depois do aniversario: continua cobrando com reajuste
            _safe_set(ws, row, 6, float(valor_com_reaj))
            _safe_set(ws, row, 7, 0)
            _safe_set(ws, row, 8, 0)
            _safe_set(ws, row, 9, float(valor_devido))
            _safe_set(ws, row, 10, float(retro_unit))

    # L18 - Total Retroativo
    L18 = linha_base + 17
    _safe_set(ws, L18, 6, "Total Retroativo a ser pago")
    _safe_set(ws, L18, 10, float(total_retro))

    # L19 - Parcelado em 12
    L19 = linha_base + 18
    _safe_set(ws, L19, 6, "Retroativo Parcelado em 12")
    _safe_set(ws, L19, 10, float(valor_parcela))


def _preencher_retroativo(ws, caso: Caso, linhas: list[LinhaSubstituicao]):
    """Gera 1 bloco de retroativo por ano com reajuste ABUSIVO.

    Limpa o template inteiro (37 linhas) e gera blocos sequenciais de 21 linhas
    cada (19 de dados + 2 de spacer). Suporta N anos abusivos.
    """
    abusivos = _anos_abusivos(linhas)
    if not abusivos:
        _limpar_intervalo(ws, 1, max(ws.max_row, 40), 1, max(ws.max_column, 10))
        _safe_set(ws, 1, 1, "Não há reajustes abusivos a retroagir neste caso.")
        return

    mes_aniv = caso.contrato.mes_aniversario
    por_ano = _linhas_por_ano(linhas)

    # Limpa todo o template (ate 37 linhas, 10 cols + folga)
    _limpar_intervalo(ws, 1, max(ws.max_row, 40), 1, max(ws.max_column, 10))

    BLOCO_TAM = 21  # 19 linhas de bloco + 2 de spacer
    # Indexa todas as linhas por (ano, mes) para lookup rapido
    por_comp: dict[tuple[int, int], LinhaSubstituicao] = {}
    for l in linhas:
        a, m = (int(x) for x in l.competencia.split("-"))
        por_comp[(a, m)] = l

    for idx, (ano, linha_aniv) in enumerate(abusivos):
        linha_base = 1 + idx * BLOCO_TAM
        # Valor base = mensalidade NORMAL imediatamente antes do aniversario.
        # Estrategia: busca mes (aniv-1) no mesmo ano, descartando anomalias (pro-rata/acerto).
        # Se nao houver, recua mes a mes ate achar um pagamento "normal".
        valor_base: Decimal | None = None
        for delta in range(1, 13):
            m = mes_aniv - delta
            a = ano
            while m <= 0:
                m += 12
                a -= 1
            l_ant = por_comp.get((a, m))
            if l_ant is None:
                continue
            if l_ant.tipo in (TipoMes.PRO_RATA, TipoMes.ACERTO):
                continue
            valor_base = l_ant.pago
            break
        if valor_base is None or valor_base == 0:
            # Ultimo fallback: deriva do proprio reajuste aplicado
            denom = (Decimal(1) + linha_aniv.reajuste_aplicado_pct)
            valor_base = (linha_aniv.pago / denom) if denom != 0 else linha_aniv.pago

        _preencher_bloco_retroativo(
            ws,
            linha_base=linha_base,
            ano=ano,
            mes_aniv=mes_aniv,
            valor_base=valor_base,
            pct_aplicado=linha_aniv.reajuste_aplicado_pct,
            pct_ans=linha_aniv.reajuste_devido_pct,
        )


def _preencher_calculo_mensalidade(ws, caso: Caso, linhas: list[LinhaSubstituicao]):
    """Tabela ano-a-ano da mensalidade aplicando ANS oficial (PEDIDO = DECISAO por default).

    L10: valor base (mensalidade inicial)
    L11..LN: ano | %ANS | valor_acumulado (PEDIDO) | ano | %ANS | valor_acumulado (DECISAO)
    """
    if not linhas:
        return

    # Limpa todas as linhas de dados (L5 em diante - inclui L5 valor total)
    _limpar_intervalo(ws, 5, max(ws.max_row, 34), 1, max(ws.max_column, 7))

    valor_base = caso.contrato.mensalidade_inicial
    if valor_base <= 0:
        valor_base = min(b.mensalidade_base for b in caso.beneficiarios)

    # Restaura cabecalhos
    _safe_set(ws, 7, 1, "BENEFICIÁRIO 1")
    _safe_set(ws, 8, 1, "PEDIDO")
    _safe_set(ws, 8, 5, "DECISÃO")
    _safe_set(ws, 9, 1, "Ano")
    _safe_set(ws, 9, 2, "Reajuste")
    _safe_set(ws, 9, 3, "Valor")
    _safe_set(ws, 9, 5, "Ano")
    _safe_set(ws, 9, 6, "Reajuste")
    _safe_set(ws, 9, 7, "Valor")

    # L10 - linha "Valor antes do reajuste anual -->"
    _safe_set(ws, 10, 1, "Valor antes do reajuste anual -->")
    _safe_set(ws, 10, 3, float(valor_base))
    _safe_set(ws, 10, 5, "Valor antes do reajuste anual -->")
    _safe_set(ws, 10, 7, float(valor_base))

    # Aplica reajustes ANS ano a ano. Corta no ultimo ano com aniversario detectado
    # ou no ano corrente, o que for menor (evita anos com %=0 poluindo).
    aniversarios = [
        int(l.competencia.split("-")[0])
        for l in linhas
        if l.tipo in (TipoMes.ANIVERSARIO_ANS, TipoMes.ANIVERSARIO_ABUSIVO, TipoMes.ANIVERSARIO_DOWNGRADE)
    ]
    ano_inicio = caso.contrato.data_assinatura.year
    ano_fim_aniv = max(aniversarios, default=date.today().year)
    ano_fim_ans = max(ANS_INDIVIDUAL.keys())
    ano_fim = min(date.today().year, ano_fim_aniv, ano_fim_ans)

    valor_pedido = valor_base
    valor_decisao = valor_base
    linha = 11
    for ano in range(ano_inicio, ano_fim + 1):
        pct_ans = ANS_INDIVIDUAL.get(ano, Decimal("0"))
        valor_pedido = valor_pedido * (Decimal(1) + pct_ans)
        valor_decisao = valor_decisao * (Decimal(1) + pct_ans)
        _safe_set(ws, linha, 1, ano)
        _safe_set(ws, linha, 2, float(pct_ans))
        _safe_set(ws, linha, 3, float(valor_pedido))
        _safe_set(ws, linha, 5, ano)
        _safe_set(ws, linha, 6, float(pct_ans))
        _safe_set(ws, linha, 7, float(valor_decisao))
        linha += 1

    # Linha final - valor da mensalidade atual
    linha_final = linha + 1
    _safe_set(ws, linha_final, 1, "VALOR MENSALIDADE DEVIDA ATUAL DE ACORDO COM A DECISÃO")
    _safe_set(ws, linha_final, 7, float(valor_decisao))

    # L5 - total geral (mesma coisa pra 1 vida)
    _safe_set(ws, 5, 1, "VALOR MENSALIDADE TOTAL DEVIDA ATUAL DE ACORDO COM A DECISÃO")
    _safe_set(ws, 5, 7, float(valor_decisao))


def _preencher_projecao(ws, caso: Caso, linhas: list[LinhaSubstituicao]):
    """Projeta operadora vs ANS ano-a-ano com diferenca e economia anual.

    L4: Nome do beneficiario
    L5: Headers
    L6+: Ano | %op | Valor pago | %ANS | Valor devido | Diferenca | Economia anual
    """
    if not linhas:
        return

    # Limpa linhas L4..fim (mantem L1-L3 que tem titulo)
    _limpar_intervalo(ws, 4, max(ws.max_row, 40), 1, max(ws.max_column, 11))

    # Restaura headers
    nome_benef = caso.beneficiarios[0].nome if caso.beneficiarios else "—"
    _safe_set(ws, 4, 1, "Nome")
    _safe_set(ws, 4, 2, nome_benef)
    _safe_set(ws, 5, 1, "Ano")
    _safe_set(ws, 5, 2, "Reajuste Aplicado")
    _safe_set(ws, 5, 3, "Valor Pago Antes Reaj. Anual")
    _safe_set(ws, 5, 4, "Reajuste Proposto")
    _safe_set(ws, 5, 5, "Valor Devido Antes Reaj. Anual")
    _safe_set(ws, 5, 6, "Diferença")
    _safe_set(ws, 5, 7, f"Economia Anual c/ Reajuste no Mês de {MESES_PT[caso.contrato.mes_aniversario - 1]}")

    # Indexa reajustes anuais do caso por ano (do tipo aniversario qualquer)
    aniv_por_ano: dict[int, LinhaSubstituicao] = {}
    for l in linhas:
        if l.tipo in (TipoMes.ANIVERSARIO_ANS, TipoMes.ANIVERSARIO_ABUSIVO, TipoMes.ANIVERSARIO_DOWNGRADE):
            ano = int(l.competencia.split("-")[0])
            aniv_por_ano[ano] = l

    valor_base = caso.contrato.mensalidade_inicial
    if valor_base <= 0:
        valor_base = min(b.mensalidade_base for b in caso.beneficiarios)

    ano_inicio = caso.contrato.data_assinatura.year
    ano_fim_aniv = max(aniv_por_ano.keys(), default=date.today().year)
    ano_fim = min(date.today().year, ano_fim_aniv)
    valor_pago = valor_base
    valor_devido = valor_base
    linha = 6
    # L6 - ano inicial (sem reajuste ainda)
    _safe_set(ws, linha, 1, ano_inicio)
    _safe_set(ws, linha, 3, float(valor_pago))
    _safe_set(ws, linha, 5, float(valor_devido))
    linha += 1
    for ano in range(ano_inicio + 1, ano_fim + 1):
        l = aniv_por_ano.get(ano)
        pct_op = l.reajuste_aplicado_pct if l else ANS_INDIVIDUAL.get(ano, Decimal("0"))
        pct_ans = ANS_INDIVIDUAL.get(ano, Decimal("0"))
        valor_pago = valor_pago * (Decimal(1) + pct_op)
        valor_devido = valor_devido * (Decimal(1) + pct_ans)
        diferenca = valor_pago - valor_devido
        economia_anual = diferenca * Decimal(12)
        _safe_set(ws, linha, 1, ano)
        _safe_set(ws, linha, 2, float(pct_op))
        _safe_set(ws, linha, 3, float(valor_pago))
        _safe_set(ws, linha, 4, float(pct_ans))
        _safe_set(ws, linha, 5, float(valor_devido))
        _safe_set(ws, linha, 6, float(diferenca))
        _safe_set(ws, linha, 7, float(economia_anual))
        linha += 1


def _ocultar_abas_calculo_nao_usadas(wb, aba_usada: str):
    """Oculta abas Calculo{N}Vida(s) não usadas sem quebrar referencias.

    Remover abas de um workbook com formulas cruzadas faz o Excel reescrever
    referencias como #REF!. Ocultar preserva a estrutura do template e evita
    que o usuário veja dados de exemplo de outras quantidades de vidas.
    """
    abas_calculo = [n for n in wb.sheetnames if n.startswith("Calculo") and "Vida" in n]
    for nome in abas_calculo:
        if nome != aba_usada:
            wb[nome].sheet_state = "hidden"
        else:
            wb[nome].sheet_state = "visible"


def _capacidade_aba_calculo(nome: str) -> int:
    if nome == "Calculo1Vida":
        return 1
    for n in range(2, 7):
        if nome == f"Calculo{n}Vidas":
            return n
    return 1


def _neutralizar_abas_calculo_ocultas(
    wb,
    aba_usada: str,
    caso: Caso,
    resultados: dict[Beneficiario, list[LinhaSubstituicao]],
):
    """Remove erros de formula das abas ocultas sem apagar a estrutura delas."""
    benefs = list(resultados.keys())
    for nome in wb.sheetnames:
        if not (nome.startswith("Calculo") and "Vida" in nome) or nome == aba_usada:
            continue
        ws = wb[nome]
        capacidade = _capacidade_aba_calculo(nome)
        _preencher_cabecalho(ws, caso, capacidade)
        for idx in range(capacidade):
            if idx < len(benefs):
                _preencher_linhas_vida(ws, benefs[idx], idx, resultados[benefs[idx]], capacidade)
            else:
                _preencher_linhas_neutras(ws, idx, capacidade)
        _preencher_total_auxiliar(ws, capacidade)


def gerar_planilha_template(
    saida: Path,
    caso: Caso,
    resultados: dict[Beneficiario, list[LinhaSubstituicao]],
    correcao: dict | None = None,
    template: Path | None = None,
) -> Path:
    """Gera planilha XLSX a partir do template oficial do escritório.

    Preserva todas as 13 abas, fórmulas e formatação. Preenche:
      - Calculo{N}Vidas com dados das vidas (remove as outras CalculoNVidas)
      - Relatorio com dados do caso
      - Honorarios com requerente/requerido
    Outras abas (% Acumulada, IndicesANSIndiv) ficam como estão.
    """
    template = template or TEMPLATE_PATH
    if not template.exists():
        raise FileNotFoundError(
            f"Template não encontrado: {template}. "
            "Execute scripts/gerar_template.py para criar."
        )

    saida.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template, saida)

    wb = openpyxl.load_workbook(saida)

    benefs = list(resultados.keys())
    n_vidas = min(len(benefs), 6)
    aba_calc = _aba_calculo_n(n_vidas)

    # Preenche cabeçalho e dados na aba CalculoNVidas
    if aba_calc in wb.sheetnames:
        ws = wb[aba_calc]
        _preencher_cabecalho(ws, caso, n_vidas)
        for i, (benef, linhas) in enumerate(resultados.items()):
            if i >= n_vidas:
                break
            _preencher_linhas_vida(ws, benef, i, linhas, n_vidas)
        _preencher_total_auxiliar(ws, n_vidas)

    # Abas ocultas tambem sao recalculadas pelo Excel; neutraliza dados antigos
    # para evitar #DIV/0! em formulas fora da aba usada.
    _neutralizar_abas_calculo_ocultas(wb, aba_calc, caso, resultados)

    # Oculta abas CalculoNVidas que não são do caso sem quebrar formulas.
    _ocultar_abas_calculo_nao_usadas(wb, aba_calc)

    # Preenche Relatorio
    if "Relatorio" in wb.sheetnames:
        _preencher_relatorio(wb["Relatorio"], caso, resultados, correcao)
        ultima_linha = max((_ultima_linha_util(ls) for ls in resultados.values()), default=LINHA_DADOS_FIM)
        _preencher_formulas_relatorio(wb["Relatorio"], caso, aba_calc, n_vidas, ultima_linha)

    # Preenche Honorarios
    if "Honorarios" in wb.sheetnames:
        total_rest = sum(
            totalizar_substituicao(ls)["restituivel_simples"]
            for _, ls in resultados.items()
        )
        _preencher_honorarios(wb["Honorarios"], caso, total_rest)

    # Preenche abas dinamicas: Retroativo, CalculoMensalidade, Projecao
    # Usa a 1a vida (multi-vida nao coberto por ora)
    if benefs:
        linhas_1a = resultados[benefs[0]]
        if "Retroativo" in wb.sheetnames:
            _preencher_retroativo(wb["Retroativo"], caso, linhas_1a)
        if "CalculoMensalidade" in wb.sheetnames:
            _preencher_calculo_mensalidade(wb["CalculoMensalidade"], caso, linhas_1a)
        if "Projecao" in wb.sheetnames:
            _preencher_projecao(wb["Projecao"], caso, linhas_1a)

    wb.save(saida)
    return saida
